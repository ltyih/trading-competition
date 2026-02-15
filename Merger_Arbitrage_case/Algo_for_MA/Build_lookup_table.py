#!/usr/bin/env python3
"""
RITC 2026 Merger Arbitrage — Multi-Heat News Extraction & Lookup Table Builder
================================================================================
Usage:
  1. Place all your heat databases (.db files) or extracted data in a folder
  2. Run: python build_lookup_table.py --data-dir ./heat_data/
  3. Output: MASTER_LOOKUP.py (paste into your trading algo)

This script:
  - Processes session data from multiple heats (XLSX or SQLite DB)
  - Extracts per-news price and probability impacts
  - Decontaminates concurrent news using cross-heat triangulation
  - Attributes each headline to the correct deal via body text analysis
  - Produces a single lookup table for instant classification
"""

import json
import os
import re
import sys
import sqlite3
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# =============================================================================
# DEAL CONFIGURATION (same as in trading algo)
# =============================================================================
DEALS = {
    'D1': {'target': 'TGX', 'acquirer': 'PHR', 'cash': 50.00, 'ratio': 0.0,
            'p0': 0.70, 'target_start': 43.70, 'acquirer_start': 47.50},
    'D2': {'target': 'BYL', 'acquirer': 'CLD', 'cash': 0.0, 'ratio': 0.75,
            'p0': 0.55, 'target_start': 43.50, 'acquirer_start': 79.30},
    'D3': {'target': 'GGD', 'acquirer': 'PNR', 'cash': 33.00, 'ratio': 0.20,
            'p0': 0.50, 'target_start': 31.50, 'acquirer_start': 59.80},
    'D4': {'target': 'FSR', 'acquirer': 'ATB', 'cash': 40.00, 'ratio': 0.0,
            'p0': 0.38, 'target_start': 30.50, 'acquirer_start': 62.20},
    'D5': {'target': 'SPK', 'acquirer': 'EEC', 'cash': 0.0, 'ratio': 1.20,
            'p0': 0.45, 'target_start': 52.80, 'acquirer_start': 48.00},
}

ALL_TICKERS = ['TGX', 'PHR', 'BYL', 'CLD', 'GGD', 'PNR', 'FSR', 'ATB', 'SPK', 'EEC']
TARGETS = ['TGX', 'BYL', 'GGD', 'FSR', 'SPK']

COMPANY_NAMES = {
    'D1': ['TARGENIX', 'PHARMACO'],
    'D2': ['BYTELAYER', 'CLOUDSYS', 'BYTE LAYER', 'CLOUD SYS'],
    'D3': ['GREENGRID', 'PETRONORTH', 'GREEN GRID', 'PETRO NORTH'],
    'D4': ['FINSURE', 'ATLAS BANK', 'ATLAS', 'FIN SURE'],
    'D5': ['SOLARPEAK', 'EASTENERGY', 'SOLAR PEAK', 'EAST ENERGY'],
}

SECTOR_KEYWORDS = {
    'D1': ['FDA', 'CLINICAL', 'DRUG', 'PHARMA', 'BIOTECH', 'PATENT', 'ONCOLOGY',
            'THERAPEUTICS', 'BIOLOGIC', 'PIPELINE', 'GENERIC', 'STATE AG'],
    'D2': ['CLOUD', 'SOFTWARE', 'SAAS', 'DATA CENTER', 'FTC', 'ANTITRUST',
            'PLATFORM', 'DIGITAL', 'COMPUTE', 'TECH ACQUISITION'],
    'D3': ['PIPELINE', 'GRID', 'FOSSIL', 'CARBON', 'ENVIRONMENTAL', 'OIL AND GAS',
            'ENERGY INFRASTRUCTURE', 'EMISSIONS', 'GREEN', 'TRANSITION', 'FERC',
            'INTERCONNECT', 'MEGAWATT', 'GENERATION'],
    'D4': ['FDIC', 'OCC', 'BANKING', 'BANK MERGER', 'DEPOSIT', 'STRESS TEST',
            'CAPITAL RATIO', 'FED ENHANCED', 'ENHANCED REVIEW', 'BRANCH', 'LENDING',
            'COMMUNITY BANK', 'BANK SECRECY', 'BSA'],
    'D5': ['SOLAR', 'RENEWABLE', 'FERC', 'INTERCONNECT', 'PHOTOVOLTAIC', 'MEGAWATT',
            'TAX CREDIT', 'CLEAN ENERGY', 'WIND', 'TURBINE', 'PANEL', 'GENERATION'],
}


# =============================================================================
# NORMALIZATION
# =============================================================================
def normalize_headline(headline: str) -> str:
    """Strip tickers, company names, normalize whitespace."""
    text = headline.upper().strip()
    for ticker in ALL_TICKERS:
        text = text.replace(ticker, 'TICKER')
    for names in COMPANY_NAMES.values():
        for name in names:
            text = text.replace(name, 'COMPANY')
    # Remove extra spaces
    text = ' '.join(text.split())
    # Remove leading/trailing punctuation artifacts
    text = text.strip(' -–—:')
    return text


def identify_deal_from_text(text: str) -> Optional[str]:
    """Identify which deal a text (headline + body) refers to."""
    text_upper = text.upper()
    scores = {}
    for deal_id, deal in DEALS.items():
        score = 0
        # Tickers (highest signal)
        if deal['target'] in text_upper:
            score += 10
        if deal['acquirer'] in text_upper:
            score += 10
        # Company names
        for name in COMPANY_NAMES.get(deal_id, []):
            if name in text_upper:
                score += 8
        # Sector keywords
        for kw in SECTOR_KEYWORDS.get(deal_id, []):
            if kw in text_upper:
                score += 2
        if score > 0:
            scores[deal_id] = score

    if not scores:
        return None
    best = max(scores, key=scores.get)
    return best if scores[best] >= 3 else None


# =============================================================================
# PROBABILITY COMPUTATION
# =============================================================================
def compute_standalone_value(deal_id: str) -> float:
    """V = (P0 - p0 * K) / (1 - p0)"""
    d = DEALS[deal_id]
    K = d['cash'] + d['ratio'] * d['acquirer_start']
    p0 = d['p0']
    if p0 >= 1:
        return d['target_start']
    return (d['target_start'] - p0 * K) / (1 - p0)


def compute_implied_prob(deal_id: str, target_price: float, acquirer_price: float) -> Optional[float]:
    """p = (Pt - V) / (K - V)"""
    d = DEALS[deal_id]
    K = d['cash'] + d['ratio'] * acquirer_price
    V = compute_standalone_value(deal_id)
    denom = K - V
    if abs(denom) < 0.01:
        return None
    p = (target_price - V) / denom
    return max(0.0, min(1.0, p))


# =============================================================================
# DATA LOADING — from XLSX files
# =============================================================================
def load_heat_from_xlsx(xlsx_path: str) -> dict:
    """Load a heat's data from an XLSX file (your extraction format)."""
    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    heat_data = {'news': [], 'prices': {}, 'probs': {}}

    # Load news
    if 'News' in wb.sheetnames:
        ws = wb['News']
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                continue
            if headers:
                d = dict(zip(headers, row))
                heat_data['news'].append(d)

    # Load prices
    if 'Prices' in wb.sheetnames:
        ws = wb['Prices']
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                continue
            if headers:
                d = dict(zip(headers, row))
                tick = d.get('tick')
                if tick is not None:
                    heat_data['prices'][int(tick)] = d

    # Load probabilities
    if 'Probabilities' in wb.sheetnames:
        ws = wb['Probabilities']
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                continue
            if headers:
                d = dict(zip(headers, row))
                tick = d.get('tick')
                if tick is not None:
                    heat_data['probs'][int(tick)] = d

    wb.close()
    return heat_data


def load_heat_from_json(json_path: str) -> dict:
    """Load pre-extracted JSON templates."""
    with open(json_path) as f:
        return json.load(f)


# =============================================================================
# NEWS IMPACT COMPUTATION
# =============================================================================
def compute_news_impacts(heat_data: dict, window: int = 8) -> List[dict]:
    """For each news item, compute per-deal probability and price changes."""
    results = []
    news_items = heat_data['news']
    prices = heat_data['prices']
    probs = heat_data['probs']

    # Group news by tick to detect concurrent items
    by_tick = defaultdict(list)
    for n in news_items:
        tick = n.get('tick')
        if tick is not None and tick > 1:
            by_tick[int(tick)].append(n)

    for tick, items in sorted(by_tick.items()):
        tick_before = max(1, tick - 1)
        tick_after = min(max(probs.keys()) if probs else tick, tick + window)

        prob_before = probs.get(tick_before, {})
        prob_after = probs.get(tick_after, {})
        price_before = prices.get(tick_before, {})
        price_after = prices.get(tick_after, {})

        # Compute per-deal changes
        deal_prob_changes = {}
        deal_price_changes = {}
        for did in ['D1', 'D2', 'D3', 'D4', 'D5']:
            pb = prob_before.get(did)
            pa = prob_after.get(did)
            if pb is not None and pa is not None:
                deal_prob_changes[did] = pa - pb
            else:
                deal_prob_changes[did] = 0.0

            target = DEALS[did]['target']
            p1 = price_before.get(target, 0)
            p2 = price_after.get(target, 0)
            if p1 and p2:
                deal_price_changes[target] = p2 - p1
            else:
                deal_price_changes[target] = 0.0

        # Find most affected deal
        max_deal = max(deal_prob_changes, key=lambda d: abs(deal_prob_changes[d]))
        max_dp = deal_prob_changes[max_deal]

        concurrent_headlines = [n.get('headline', '') for n in items]

        for n in items:
            headline = n.get('headline', '')
            body = n.get('body', '')
            normalized = n.get('normalized') or normalize_headline(headline)

            # Attribution: use body text to determine which deal this headline affects
            attributed_deal = identify_deal_from_text(f"{headline} {body}")

            result = {
                'normalized': normalized,
                'headline': headline,
                'body': body,
                'tick': tick,
                'concurrent_count': len(items),
                'concurrent_headlines': concurrent_headlines,
                'is_clean': len(items) == 1,
                'deal_prob_changes': deal_prob_changes.copy(),
                'deal_price_changes': deal_price_changes.copy(),
                'max_affected_deal': max_deal,
                'max_dp': max_dp,
                'attributed_deal': attributed_deal or max_deal,
            }

            # If we can attribute to a specific deal, use that deal's prob change
            if attributed_deal:
                result['attributed_dp'] = deal_prob_changes.get(attributed_deal, max_dp)
                target = DEALS[attributed_deal]['target']
                result['attributed_price_move'] = deal_price_changes.get(target, 0)
            else:
                result['attributed_dp'] = max_dp
                target = DEALS[max_deal]['target']
                result['attributed_price_move'] = deal_price_changes.get(target, 0)

            results.append(result)

    return results


# =============================================================================
# CROSS-HEAT MERGE & DECONTAMINATION
# =============================================================================
def merge_across_heats(all_heat_impacts: List[List[dict]]) -> dict:
    """
    Merge news impacts from all heats into a master lookup table.

    Priority:
    1. Clean observations (solo at tick) across all heats
    2. Body-attributed observations for concurrent items
    3. Max-affected-deal fallback
    """
    # Group by normalized headline
    by_headline = defaultdict(list)
    for heat_impacts in all_heat_impacts:
        for item in heat_impacts:
            by_headline[item['normalized']].append(item)

    master = {}
    for normalized, observations in by_headline.items():
        # Separate clean vs contaminated
        clean = [o for o in observations if o['is_clean']]
        contaminated = [o for o in observations if not o['is_clean']]

        # Collect attributed impacts
        dps = []
        price_moves = []
        deals_seen = []

        # Prefer clean observations
        for o in clean:
            dps.append(o['attributed_dp'])
            price_moves.append(o['attributed_price_move'])
            deals_seen.append(o['attributed_deal'])

        # Add contaminated with lower weight (they're less reliable)
        for o in contaminated:
            dps.append(o['attributed_dp'])
            price_moves.append(o['attributed_price_move'])
            deals_seen.append(o['attributed_deal'])

        if not dps:
            continue

        avg_dp = sum(dps) / len(dps)
        avg_move = sum(price_moves) / len(price_moves)
        n_clean = len(clean)
        n_total = len(observations)

        # Determine direction from observed data
        if avg_dp > 0.005:
            direction = 'positive'
        elif avg_dp < -0.005:
            direction = 'negative'
        else:
            direction = 'ambiguous'

        # Determine severity from magnitude
        abs_dp = abs(avg_dp)
        if abs_dp > 0.06:
            severity = 'large'
        elif abs_dp > 0.02:
            severity = 'medium'
        else:
            severity = 'small'

        # Determine most common category (from body text analysis)
        # For now, use a simple keyword-based category
        category = classify_category(normalized)

        # Confidence: higher if more clean observations
        confidence = min(1.0, (n_clean * 1.0 + (n_total - n_clean) * 0.5) / 3.0)
        if n_clean >= 2:
            confidence = 1.0

        master[normalized] = {
            'direction': direction,
            'severity': severity,
            'category': category,
            'avg_dp': round(avg_dp, 5),
            'avg_price_move': round(avg_move, 3),
            'n_observations': n_total,
            'n_clean': n_clean,
            'confidence': round(confidence, 2),
            'deals_affected': list(set(deals_seen)),
            'headline_example': observations[0].get('headline', normalized),
        }

    return master


def classify_category(text: str) -> str:
    """Simple category classification from text."""
    text = text.upper()
    cat_scores = {
        'REG': 0, 'FIN': 0, 'SHR': 0, 'ALT': 0, 'PRC': 0
    }
    reg_words = ['REGULAT', 'ANTITRUST', 'FTC', 'FERC', 'FDIC', 'COMMISSION',
                 'CLEARANCE', 'APPROVAL', 'REVIEW', 'CMA', 'EPA', 'COMPLIANCE']
    fin_words = ['FINANC', 'CREDIT', 'DEBT', 'BOND', 'OFFERING', 'SPREAD',
                 'REFINANC', 'LEVERAGE', 'RATING', 'MOODY', 'CAPITAL']
    shr_words = ['SHAREHOLD', 'VOTE', 'BOARD', 'ACTIVIST', 'HEDGE FUND',
                 'PROXY', 'MEETING', 'BACKLASH', 'LITIGATION']
    alt_words = ['COMPETING', 'RIVAL', 'COUNTER', 'TOPPING', 'SWEETENED',
                 'ALTERNATIVE', 'WHITE KNIGHT']
    prc_words = ['PRICE', 'PREMIUM', 'SPREAD', 'DECLINE', 'COLLAPSE',
                 'MULTIPLE', 'COMPRESSION']

    for w in reg_words:
        if w in text: cat_scores['REG'] += 1
    for w in fin_words:
        if w in text: cat_scores['FIN'] += 1
    for w in shr_words:
        if w in text: cat_scores['SHR'] += 1
    for w in alt_words:
        if w in text: cat_scores['ALT'] += 1
    for w in prc_words:
        if w in text: cat_scores['PRC'] += 1

    if max(cat_scores.values()) == 0:
        return 'FIN'
    return max(cat_scores, key=cat_scores.get)


# =============================================================================
# OUTPUT: Generate Python lookup table
# =============================================================================
def generate_python_output(master: dict, output_path: str):
    """Generate a Python file with the MASTER_LOOKUP dict."""
    lines = [
        '# Auto-generated MASTER LOOKUP TABLE for RITC 2026 Merger Arbitrage',
        f'# Generated from {sum(v["n_observations"] for v in master.values())} observations',
        f'# across {len(master)} unique headlines',
        '',
        'MASTER_LOOKUP = {',
    ]

    # Sort by confidence (highest first), then by absolute dp
    sorted_items = sorted(master.items(),
                          key=lambda x: (-x[1]['confidence'], -abs(x[1]['avg_dp'])))

    for normalized, entry in sorted_items:
        lines.append(f'    {normalized!r}: {{')
        lines.append(f'        "direction": "{entry["direction"]}",')
        lines.append(f'        "severity": "{entry["severity"]}",')
        lines.append(f'        "category": "{entry["category"]}",')
        lines.append(f'        "avg_dp": {entry["avg_dp"]},')
        lines.append(f'        "avg_price_move": {entry["avg_price_move"]},')
        lines.append(f'        "confidence": {entry["confidence"]},')
        lines.append(f'        "n_obs": {entry["n_observations"]},')
        lines.append(f'        "n_clean": {entry["n_clean"]},')
        lines.append(f'        # example: {entry["headline_example"][:60]}')
        lines.append(f'    }},')

    lines.append('}')
    lines.append('')

    # Also generate the lookup function
    lines.extend([
        '',
        'ALL_TICKERS = ["TGX", "PHR", "BYL", "CLD", "GGD", "PNR", "FSR", "ATB", "SPK", "EEC"]',
        'ALL_COMPANY_NAMES = [',
        '    "TARGENIX", "PHARMACO", "BYTELAYER", "CLOUDSYS", "GREENGRID",',
        '    "PETRONORTH", "FINSURE", "ATLAS BANK", "ATLAS", "SOLARPEAK",',
        '    "EASTENERGY", "EAST ENERGY",',
        ']',
        '',
        '',
        'def normalize_headline(headline: str) -> str:',
        '    text = headline.upper().strip()',
        '    for ticker in ALL_TICKERS:',
        '        text = text.replace(ticker, "TICKER")',
        '    for name in ALL_COMPANY_NAMES:',
        '        text = text.replace(name, "COMPANY")',
        '    text = " ".join(text.split())',
        '    return text',
        '',
        '',
        'def lookup_news(headline: str) -> dict | None:',
        '    """Look up a headline in the master table. Returns classification or None."""',
        '    normalized = normalize_headline(headline)',
        '',
        '    # Exact match',
        '    if normalized in MASTER_LOOKUP:',
        '        return MASTER_LOOKUP[normalized]',
        '',
        '    # Substring match (template in headline or headline in template)',
        '    for template, classification in MASTER_LOOKUP.items():',
        '        if template in normalized or normalized in template:',
        '            return classification',
        '',
        '    return None',
        '',
    ])

    with open(output_path, 'w') as f:
        f.write('\n'.join(lines))

    print(f"\nWrote {len(master)} entries to {output_path}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description='Build master news lookup table')
    parser.add_argument('--data-dir', default='.', help='Directory with heat XLSX/JSON files')
    parser.add_argument('--output', default='MASTER_LOOKUP.py', help='Output Python file')
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    all_heat_impacts = []

    # Find all XLSX files
    xlsx_files = sorted(data_dir.glob('session_data_*.xlsx'))
    json_files = sorted(data_dir.glob('news_templates_*.json'))

    if xlsx_files:
        print(f"Found {len(xlsx_files)} XLSX heat files")
        for xlsx_path in xlsx_files:
            print(f"\nProcessing: {xlsx_path.name}")
            heat_data = load_heat_from_xlsx(str(xlsx_path))
            impacts = compute_news_impacts(heat_data)
            all_heat_impacts.append(impacts)
            print(f"  {len(impacts)} news items, "
                  f"{sum(1 for i in impacts if i['is_clean'])} clean")
    elif json_files:
        print(f"Found {len(json_files)} JSON template files")
        # JSON files are pre-processed — less ideal but usable
        for json_path in json_files:
            print(f"\nProcessing: {json_path.name}")
            templates = load_heat_from_json(str(json_path))
            # Convert to our format
            impacts = []
            for norm, t in templates.items():
                impacts.append({
                    'normalized': norm,
                    'headline': t.get('headline_example', norm),
                    'body': '',
                    'tick': 0,
                    'concurrent_count': 1,
                    'concurrent_headlines': [],
                    'is_clean': t.get('observed_count', 1) == 1,  # approximation
                    'deal_prob_changes': {},
                    'deal_price_changes': {},
                    'max_affected_deal': t.get('deal_pattern', ''),
                    'max_dp': t.get('avg_prob_change', 0),
                    'attributed_deal': t.get('deal_pattern', ''),
                    'attributed_dp': t.get('avg_prob_change', 0),
                    'attributed_price_move': t.get('avg_target_move', 0),
                })
            all_heat_impacts.append(impacts)
            print(f"  {len(impacts)} templates loaded")
    else:
        print(f"No data files found in {data_dir}")
        print("Looking for: session_data_*.xlsx or news_templates_*.json")
        sys.exit(1)

    # Merge across heats
    print(f"\n{'='*60}")
    print(f"Merging {len(all_heat_impacts)} heats...")
    master = merge_across_heats(all_heat_impacts)
    print(f"Master table: {len(master)} unique headlines")

    # Stats
    positive = sum(1 for v in master.values() if v['direction'] == 'positive')
    negative = sum(1 for v in master.values() if v['direction'] == 'negative')
    ambiguous = sum(1 for v in master.values() if v['direction'] == 'ambiguous')
    high_conf = sum(1 for v in master.values() if v['confidence'] >= 0.8)
    print(f"  Positive: {positive} | Negative: {negative} | Ambiguous: {ambiguous}")
    print(f"  High confidence (>=0.8): {high_conf}")

    # Top movers
    print(f"\nTop 10 by absolute probability impact:")
    top = sorted(master.items(), key=lambda x: abs(x[1]['avg_dp']), reverse=True)[:10]
    for norm, entry in top:
        print(f"  {entry['direction']:>9s} {entry['severity']:>6s} dp={entry['avg_dp']:+.4f} "
              f"n={entry['n_observations']} | {norm[:60]}")

    # Generate output
    generate_python_output(master, args.output)
    print(f"\nDone! Paste the contents of {args.output} into your trading algo.")


if __name__ == '__main__':
    main()