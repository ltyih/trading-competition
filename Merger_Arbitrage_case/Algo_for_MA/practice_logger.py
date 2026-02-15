# -*- coding: utf-8 -*-
"""
PRACTICE SESSION LOGGER - Captures EVERYTHING for strategy optimization
========================================================================
Run this during every practice session. It records:

1. TICK DATA     - Every tick: all 10 security prices, bid/ask, volume
2. NEWS          - Every headline + body + the tick it appeared
3. POSITIONS     - Your positions and P&L at every snapshot
4. TRADES        - Every order you place (from /orders endpoint)
5. TENDERS       - Every tender offer that appears
6. PROBABILITIES - Market-implied deal completion probabilities
7. NEWS IMPACT   - Price changes BEFORE and AFTER each news item

At session end, it generates:
- SQLite database with all raw data
- news_templates.json - auto-built lookup table for merger_arb_trader.py
- session_report.txt  - human-readable summary
- session_data.xlsx   - Excel workbook with all tables (if openpyxl installed)

Usage:
    python practice_logger.py
    python practice_logger.py --api-key YOUR_KEY --port 9999
"""

import requests
import sqlite3
import json
import time
import sys
import os
import argparse
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Optional, Any
from concurrent.futures import ThreadPoolExecutor

# =============================================================================
# CONFIGURATION
# =============================================================================
DEFAULT_API_KEY = 'AJDSYHVC'
DEFAULT_BASE = 'http://localhost:10000/v1'

DEALS = {
    'D1': {'target': 'TGX', 'acquirer': 'PHR', 'cash': 50.00, 'ratio': 0.0,
            'p0': 0.70, 't_start': 43.70, 'a_start': 47.50},
    'D2': {'target': 'BYL', 'acquirer': 'CLD', 'cash': 0.0, 'ratio': 0.75,
            'p0': 0.55, 't_start': 43.50, 'a_start': 79.30},
    'D3': {'target': 'GGD', 'acquirer': 'PNR', 'cash': 33.00, 'ratio': 0.20,
            'p0': 0.50, 't_start': 31.50, 'a_start': 59.80},
    'D4': {'target': 'FSR', 'acquirer': 'ATB', 'cash': 40.00, 'ratio': 0.0,
            'p0': 0.38, 't_start': 30.50, 'a_start': 62.20},
    'D5': {'target': 'SPK', 'acquirer': 'EEC', 'cash': 0.0, 'ratio': 1.20,
            'p0': 0.45, 't_start': 52.80, 'a_start': 48.00},
}

ALL_TICKERS = ['TGX', 'PHR', 'BYL', 'CLD', 'GGD', 'PNR', 'FSR', 'ATB', 'SPK', 'EEC']
TARGETS = {d['target']: did for did, d in DEALS.items()}
ACQUIRERS = {d['acquirer']: did for did, d in DEALS.items()}

# How many ticks before/after news to measure price impact
NEWS_IMPACT_WINDOW_BEFORE = 3
NEWS_IMPACT_WINDOW_AFTER = 10

POLL_INTERVAL = 0.3  # seconds between data captures


# =============================================================================
# PROBABILITY MATH
# =============================================================================
standalone_values = {}
for did, d in DEALS.items():
    K0 = d['cash'] + d['ratio'] * d['a_start']
    V = (d['t_start'] - d['p0'] * K0) / (1 - d['p0'])
    standalone_values[did] = V


def deal_value(did, ap):
    d = DEALS[did]
    return d['cash'] + d['ratio'] * ap


def implied_prob(did, tp, ap):
    K = deal_value(did, ap)
    V = standalone_values[did]
    denom = K - V
    if abs(denom) < 0.01:
        return None
    return max(0.0, min(1.0, (tp - V) / denom))


# =============================================================================
# DATABASE SETUP
# =============================================================================
def create_database(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS ticks (
        tick INTEGER,
        timestamp TEXT,
        nlv REAL,
        gross_position REAL,
        net_position REAL,
        PRIMARY KEY (tick)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS prices (
        tick INTEGER,
        ticker TEXT,
        bid REAL,
        ask REAL,
        last REAL,
        mid REAL,
        volume INTEGER,
        position INTEGER,
        PRIMARY KEY (tick, ticker)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS probabilities (
        tick INTEGER,
        deal_id TEXT,
        market_implied_prob REAL,
        deal_value REAL,
        standalone_value REAL,
        spread REAL,
        target_price REAL,
        acquirer_price REAL,
        PRIMARY KEY (tick, deal_id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS news (
        news_id INTEGER PRIMARY KEY,
        tick INTEGER,
        headline TEXT,
        body TEXT,
        headline_normalized TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS news_impact (
        news_id INTEGER,
        ticker TEXT,
        price_before REAL,
        price_after REAL,
        price_change REAL,
        price_change_pct REAL,
        prob_before REAL,
        prob_after REAL,
        prob_change REAL,
        deal_id TEXT,
        is_target INTEGER,
        PRIMARY KEY (news_id, ticker)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS orders (
        order_id INTEGER PRIMARY KEY,
        tick INTEGER,
        ticker TEXT,
        type TEXT,
        action TEXT,
        quantity INTEGER,
        price REAL,
        status TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS tenders (
        tender_id INTEGER PRIMARY KEY,
        tick_seen INTEGER,
        ticker TEXT,
        action TEXT,
        price REAL,
        quantity INTEGER
    )''')

    conn.commit()
    return conn


# =============================================================================
# NEWS NORMALIZER (same as merger_arb_trader.py)
# =============================================================================
ALL_COMPANY_NAMES = [
    'TARGENIX', 'PHARMACO', 'BYTELAYER', 'CLOUDSYS', 'GREENGRID',
    'PETRONORTH', 'FINSURE', 'ATLAS BANK', 'ATLAS', 'SOLARPEAK',
    'EASTENERGY', 'EAST ENERGY',
]


def normalize_headline(headline: str) -> str:
    text = headline.upper().strip()
    for ticker in ALL_TICKERS:
        text = text.replace(ticker, 'TICKER')
    for name in ALL_COMPANY_NAMES:
        text = text.replace(name, 'COMPANY')
    text = ' '.join(text.split())
    return text


def identify_deal_from_text(text: str) -> Optional[str]:
    """Identify which deal a news item relates to."""
    text_upper = text.upper()
    for ticker in ALL_TICKERS:
        if ticker in text_upper:
            if ticker in TARGETS:
                return TARGETS[ticker]
            if ticker in ACQUIRERS:
                return ACQUIRERS[ticker]
    for did, d in DEALS.items():
        for name in ALL_COMPANY_NAMES:
            if name in text_upper:
                # Check if name belongs to this deal
                if d['target'] in ['TGX', 'BYL', 'GGD', 'FSR', 'SPK']:
                    pass  # checked above via tickers
    return None


# =============================================================================
# MAIN LOGGER
# =============================================================================
class PracticeLogger:

    def __init__(self, api_key: str, base_url: str, output_dir: str):
        self.api_key = api_key
        self.base_url = base_url
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        self.session = requests.Session()
        self.session.headers.update({'X-API-Key': api_key})

        # Cumulative templates across ALL heats (the real prize)
        self.cumulative_templates_path = os.path.join(output_dir, 'news_templates_ALL.json')
        self.cumulative_templates: Dict[str, Any] = {}
        if os.path.exists(self.cumulative_templates_path):
            with open(self.cumulative_templates_path, 'r') as f:
                self.cumulative_templates = json.load(f)
            print(f"  Loaded {len(self.cumulative_templates)} existing templates")

        self.heat_count = 0
        self.running = True

        # Per-heat state (reset each heat)
        self.conn: Optional[sqlite3.Connection] = None
        self.db_path = ''
        self._reset_heat_state()

    def _reset_heat_state(self):
        """Reset all per-heat tracking. Called between heats."""
        self.tick = 0
        self.heat_ts = ''
        self.heat_label = ''
        self.last_news_id = 0
        self.last_order_check = 0
        self.seen_order_ids = set()
        self.seen_tender_ids = set()
        self.price_history = {}
        self.prob_history = {}
        self.pending_news = []
        self.all_news = []
        self.total_news = 0

    def _start_new_heat_db(self):
        """Create a fresh database for this heat."""
        if self.conn:
            self.conn.close()
        self.heat_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.heat_label = f'heat{self.heat_count + 1}_{self.heat_ts}'
        self.db_path = os.path.join(self.output_dir,
                                     f'{self.heat_label}.db')
        self.conn = create_database(self.db_path)
        print(f"\n  New DB: {self.db_path}")

    def _end_heat(self):
        """Finalize the current heat: compute impacts, generate reports, reset."""
        if not self.conn:
            return

        print(f"\n{'=' * 80}")
        print(f"  HEAT {self.heat_count + 1} ENDED - Generating reports...")
        print(f"{'=' * 80}")

        self._compute_all_pending_impacts()
        self.conn.commit()
        self._generate_reports()

        # Merge this heat's templates into the cumulative file
        self._merge_cumulative_templates()

        self.conn.close()
        self.conn = None
        self.heat_count += 1

        print(f"\n  Heat {self.heat_count} complete. "
              f"Cumulative templates: {len(self.cumulative_templates)}")
        print(f"  Waiting for next heat...\n")

        self._reset_heat_state()

    def _merge_cumulative_templates(self):
        """Merge this heat's templates into the cumulative cross-session file."""
        # Load this heat's templates
        heat_json = os.path.join(self.output_dir, f'news_templates_{self.heat_label}.json')
        if not os.path.exists(heat_json):
            return

        with open(heat_json, 'r') as f:
            heat_templates = json.load(f)

        for norm, entry in heat_templates.items():
            if norm in self.cumulative_templates:
                # Merge: update counts and running averages
                existing = self.cumulative_templates[norm]
                old_count = existing.get('observed_count', 1)
                new_count = old_count + entry.get('observed_count', 1)

                old_move = existing.get('avg_target_move', 0)
                new_move = entry.get('avg_target_move', 0)
                merged_move = (old_move * old_count + new_move * entry.get('observed_count', 1)) / new_count

                old_prob = existing.get('avg_prob_change', 0)
                new_prob = entry.get('avg_prob_change', 0)
                merged_prob = (old_prob * old_count + new_prob * entry.get('observed_count', 1)) / new_count

                existing['observed_count'] = new_count
                existing['avg_target_move'] = merged_move
                existing['avg_prob_change'] = merged_prob
                # Keep the direction from whichever has more observations
                if entry.get('observed_count', 1) > old_count:
                    existing['direction'] = entry['direction']
                    existing['severity'] = entry['severity']
                    existing['category'] = entry['category']
            else:
                self.cumulative_templates[norm] = dict(entry)

        # Save cumulative file
        with open(self.cumulative_templates_path, 'w') as f:
            json.dump(self.cumulative_templates, f, indent=2, sort_keys=True)

        # Also generate the cumulative Python snippet
        py_path = os.path.join(self.output_dir, 'news_templates_ALL_snippet.py')
        with open(py_path, 'w') as f:
            f.write("# CUMULATIVE templates from ALL practice sessions\n")
            f.write(f"# Total templates: {len(self.cumulative_templates)}\n")
            f.write(f"# Last updated: {datetime.now().isoformat()}\n")
            f.write("# Copy into NEWS_TEMPLATES in merger_arb_trader.py\n\n")
            f.write("NEW_TEMPLATES = {\n")
            for norm, entry in sorted(self.cumulative_templates.items()):
                d = entry['direction']
                s = entry['severity']
                cat = entry['category']
                conf = entry.get('confidence', 1.0)
                move = entry.get('avg_target_move', 0)
                prob = entry.get('avg_prob_change', 0)
                count = entry.get('observed_count', 1)
                f.write(f"    '{norm}': {{\n")
                f.write(f"        'direction': '{d}', 'severity': '{s}', "
                        f"'category': '{cat}', 'confidence': {conf},\n")
                f.write(f"        # seen {count}x across all sessions, "
                        f"avg move=${move:+.3f}, avg dp={prob:+.4f}\n")
                f.write(f"    }},\n")
            f.write("}\n")

        print(f"    Cumulative: {len(self.cumulative_templates)} templates -> "
              f"{self.cumulative_templates_path}")

    def run(self):
        print("=" * 80)
        print("  PRACTICE SESSION LOGGER - CONTINUOUS MODE")
        print("  Runs across multiple heats, auto-generates reports per heat")
        print("  Builds cumulative word bank across ALL sessions")
        print("=" * 80)
        print(f"  API: {self.base_url}")
        print(f"  Output: {self.output_dir}")
        print(f"  Poll: {POLL_INTERVAL}s")
        print(f"  Existing templates: {len(self.cumulative_templates)}")
        print("=" * 80)
        print()

        # Wait for connection
        print("Waiting for RIT client...")
        while self.running:
            try:
                r = self.session.get(f'{self.base_url}/case', timeout=5)
                if r.ok:
                    print(f"Connected! Status: {r.json().get('status')}")
                    break
            except:
                pass
            time.sleep(2)

        last_tick = -1
        heat_active = False

        while self.running:
            try:
                # Get case status
                r = self.session.get(f'{self.base_url}/case', timeout=5)
                if not r.ok:
                    time.sleep(0.5)
                    continue
                case = r.json()
                self.tick = case.get('tick', 0)
                status = case.get('status', '')

                if status not in ('ACTIVE', 'RUNNING'):
                    if heat_active and last_tick > 0:
                        # Heat just ended - generate reports
                        self._end_heat()
                        heat_active = False
                        last_tick = -1
                    time.sleep(1)
                    continue

                # Heat is active
                if not heat_active:
                    # New heat starting - create fresh DB
                    self._start_new_heat_db()
                    heat_active = True
                    last_tick = -1
                    print(f"\n  Heat {self.heat_count + 1} started at tick {self.tick}")

                # Only capture if tick changed
                if self.tick == last_tick:
                    time.sleep(0.05)
                    continue
                last_tick = self.tick

                # Parallel fetch everything
                self._capture_all()

                # Progress indicator
                if self.tick % 30 == 0:
                    self._print_status()

                time.sleep(POLL_INTERVAL)

            except KeyboardInterrupt:
                print("\n\nStopping logger...")
                self.running = False
            except Exception as e:
                print(f"Error: {e}")
                time.sleep(1)

        # Final processing if a heat was in progress
        if heat_active and self.conn:
            self._end_heat()

        print(f"\nLogger finished. {self.heat_count} heats recorded.")
        print(f"Cumulative templates: {len(self.cumulative_templates)}")
        print(f"All data in: {self.output_dir}/")

    def _capture_all(self):
        """Capture all data sources in parallel."""
        with ThreadPoolExecutor(max_workers=4) as pool:
            f_sec = pool.submit(self._safe_get, f'{self.base_url}/securities')
            f_news = pool.submit(self._safe_get, f'{self.base_url}/news',
                                 {'since': self.last_news_id, 'limit': 100})
            f_trader = pool.submit(self._safe_get, f'{self.base_url}/trader')
            f_tenders = pool.submit(self._safe_get, f'{self.base_url}/tenders')

            sec_data = f_sec.result()
            news_data = f_news.result()
            trader_data = f_trader.result()
            tender_data = f_tenders.result()

        # Process securities (prices + positions)
        prices = {}
        positions = {}
        if sec_data:
            for sec in sec_data:
                ticker = sec.get('ticker', '')
                bid = sec.get('bid', 0) or 0
                ask = sec.get('ask', 0) or 0
                last = sec.get('last', 0) or 0
                mid = round((bid + ask) / 2, 4) if bid and ask else last
                vol = sec.get('volume', 0) or 0
                pos = sec.get('position', 0) or 0

                prices[ticker] = mid
                positions[ticker] = pos

                self.conn.execute(
                    'INSERT OR REPLACE INTO prices VALUES (?,?,?,?,?,?,?,?)',
                    (self.tick, ticker, bid, ask, last, mid, vol, pos))

        # Store price snapshot for news impact calculation
        if prices:
            self.price_history[self.tick] = dict(prices)

        # Process trader (NLV)
        nlv = 0
        if trader_data:
            nlv = trader_data.get('nlv', 0) or 0

        gross = sum(abs(v) for v in positions.values())
        net = sum(v for v in positions.values())

        self.conn.execute(
            'INSERT OR REPLACE INTO ticks VALUES (?,?,?,?,?)',
            (self.tick, datetime.now().isoformat(), nlv, gross, net))

        # Process probabilities
        prob_snap = {}
        for did, deal in DEALS.items():
            tp = prices.get(deal['target'], 0)
            ap = prices.get(deal['acquirer'], 0)
            if tp > 0 and ap > 0:
                K = deal_value(did, ap)
                V = standalone_values[did]
                mp = implied_prob(did, tp, ap)
                spread = K - tp
                if mp is not None:
                    prob_snap[did] = mp
                    self.conn.execute(
                        'INSERT OR REPLACE INTO probabilities VALUES (?,?,?,?,?,?,?,?)',
                        (self.tick, did, mp, K, V, spread, tp, ap))

        self.prob_history[self.tick] = prob_snap

        # Process news
        if news_data:
            for n in sorted(news_data, key=lambda x: x.get('news_id', 0)):
                nid = n.get('news_id', 0)
                if nid <= self.last_news_id:
                    continue

                headline = n.get('headline', '')
                body = n.get('body', '')
                normalized = normalize_headline(headline)

                self.conn.execute(
                    'INSERT OR REPLACE INTO news VALUES (?,?,?,?,?)',
                    (nid, self.tick, headline, body, normalized))

                deal_id = identify_deal_from_text(f"{headline} {body}")
                news_entry = {
                    'news_id': nid,
                    'tick': self.tick,
                    'headline': headline,
                    'body': body,
                    'normalized': normalized,
                    'deal_id': deal_id,
                }
                self.all_news.append(news_entry)
                self.pending_news.append(news_entry)
                self.total_news += 1
                self.last_news_id = nid

                print(f"\n  *** NEWS #{nid} tick {self.tick}: {headline}")
                if body:
                    print(f"      BODY: {body[:200]}")

        # Process tenders
        if tender_data:
            for t in tender_data:
                tid = t.get('tender_id', 0)
                if tid not in self.seen_tender_ids:
                    self.seen_tender_ids.add(tid)
                    self.conn.execute(
                        'INSERT OR REPLACE INTO tenders VALUES (?,?,?,?,?,?)',
                        (tid, self.tick, t.get('ticker', ''), t.get('action', ''),
                         t.get('price', 0), t.get('quantity', 0)))
                    print(f"  TENDER #{tid}: {t.get('action')} {t.get('quantity')} "
                          f"{t.get('ticker')} @ ${t.get('price', 0):.2f}")

        # Check orders periodically
        if self.tick - self.last_order_check >= 5:
            self._capture_orders()
            self.last_order_check = self.tick

        # Compute impacts for news that now have enough post-news ticks
        self._compute_ready_impacts()

        self.conn.commit()

    def _capture_orders(self):
        """Capture any new orders from the orders endpoint."""
        for status_filter in ['TRANSACTED', 'OPEN', 'CANCELLED']:
            data = self._safe_get(f'{self.base_url}/orders',
                                  {'status': status_filter})
            if data:
                for o in data:
                    oid = o.get('order_id', 0)
                    if oid not in self.seen_order_ids:
                        self.seen_order_ids.add(oid)
                        self.conn.execute(
                            'INSERT OR REPLACE INTO orders VALUES (?,?,?,?,?,?,?,?)',
                            (oid, o.get('tick', self.tick), o.get('ticker', ''),
                             o.get('type', ''), o.get('action', ''),
                             o.get('quantity', 0), o.get('price', 0),
                             o.get('status', status_filter)))

    def _compute_ready_impacts(self):
        """Compute price impact for news items that now have enough after-data."""
        still_pending = []
        for news in self.pending_news:
            news_tick = news['tick']
            after_tick = news_tick + NEWS_IMPACT_WINDOW_AFTER

            if self.tick < after_tick:
                still_pending.append(news)
                continue

            self._compute_single_impact(news)

        self.pending_news = still_pending

    def _compute_all_pending_impacts(self):
        """Compute impact for ALL remaining pending news (end of heat)."""
        for news in self.pending_news:
            self._compute_single_impact(news)
        self.pending_news = []

    def _compute_single_impact(self, news: Dict):
        """Compute the price/probability impact of a single news item."""
        news_tick = news['tick']
        nid = news['news_id']

        # Find the closest price snapshot BEFORE the news
        before_tick = None
        for t in range(news_tick - 1, news_tick - NEWS_IMPACT_WINDOW_BEFORE - 1, -1):
            if t in self.price_history:
                before_tick = t
                break
        if before_tick is None:
            before_tick = news_tick  # fallback to same tick

        # Find the closest price snapshot AFTER the news (target: +10 ticks)
        after_tick = None
        for t in range(news_tick + NEWS_IMPACT_WINDOW_AFTER,
                       news_tick, -1):
            if t in self.price_history:
                after_tick = t
                break
        if after_tick is None:
            # Use the latest available tick
            available = [t for t in self.price_history if t > news_tick]
            after_tick = min(available) if available else news_tick

        before_prices = self.price_history.get(before_tick, {})
        after_prices = self.price_history.get(after_tick, {})
        before_probs = self.prob_history.get(before_tick, {})
        after_probs = self.prob_history.get(after_tick, {})

        for ticker in ALL_TICKERS:
            pb = before_prices.get(ticker, 0)
            pa = after_prices.get(ticker, 0)
            if pb <= 0 or pa <= 0:
                continue

            change = pa - pb
            change_pct = (change / pb) * 100 if pb > 0 else 0

            # Determine deal and if this ticker is target or acquirer
            deal_id = None
            is_target = 0
            for did, deal in DEALS.items():
                if ticker == deal['target']:
                    deal_id = did
                    is_target = 1
                    break
                elif ticker == deal['acquirer']:
                    deal_id = did
                    is_target = 0
                    break

            prob_before = before_probs.get(deal_id, 0) if deal_id else 0
            prob_after = after_probs.get(deal_id, 0) if deal_id else 0
            prob_change = prob_after - prob_before

            self.conn.execute(
                'INSERT OR REPLACE INTO news_impact VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (nid, ticker, pb, pa, change, change_pct,
                 prob_before, prob_after, prob_change, deal_id, is_target))

    def _safe_get(self, url, params=None):
        """Safe API GET with timeout handling."""
        try:
            r = self.session.get(url, params=params, timeout=5)
            if r.ok:
                return r.json()
        except:
            pass
        return None

    def _print_status(self):
        """Print periodic status."""
        # Quick NLV fetch
        trader = self._safe_get(f'{self.base_url}/trader')
        nlv = trader.get('nlv', 0) if trader else 0

        probs = []
        for did in ['D1', 'D2', 'D3', 'D4', 'D5']:
            latest = self.prob_history.get(self.tick, {}).get(did)
            if latest is not None:
                probs.append(f"{did}={latest:.0%}")
            else:
                probs.append(f"{did}=?")

        print(f"\n  [Tick {self.tick:3d}] NLV=${nlv:,.0f} | "
              f"News={self.total_news} | {' '.join(probs)}")

    # =========================================================================
    # REPORT GENERATION
    # =========================================================================
    def _generate_reports(self):
        """Generate all analysis outputs."""
        print("\n" + "=" * 80)
        print("  GENERATING REPORTS")
        print("=" * 80)

        self._generate_news_templates()
        self._generate_text_report()
        self._generate_excel_report()

        print(f"\n  All files saved to: {self.output_dir}/")
        print("=" * 80)

    def _generate_news_templates(self):
        """Build the news_templates.json lookup table from observed data."""
        print("  Building news_templates.json ...")

        c = self.conn.cursor()

        # For each news item, find which target moved the most
        templates = {}

        c.execute('''
            SELECT n.news_id, n.tick, n.headline, n.body, n.headline_normalized,
                   ni.ticker, ni.price_change, ni.price_change_pct,
                   ni.prob_before, ni.prob_after, ni.prob_change,
                   ni.deal_id, ni.is_target
            FROM news n
            LEFT JOIN news_impact ni ON n.news_id = ni.news_id AND ni.is_target = 1
            ORDER BY n.news_id, abs(ni.price_change) DESC
        ''')

        news_impacts = defaultdict(list)
        news_info = {}
        for row in c.fetchall():
            nid = row[0]
            if nid not in news_info:
                news_info[nid] = {
                    'tick': row[1], 'headline': row[2], 'body': row[3],
                    'normalized': row[4],
                }
            if row[5]:  # has impact data
                news_impacts[nid].append({
                    'ticker': row[5], 'price_change': row[6],
                    'price_change_pct': row[7],
                    'prob_before': row[8], 'prob_after': row[9],
                    'prob_change': row[10], 'deal_id': row[11],
                    'is_target': row[12],
                })

        for nid, info in news_info.items():
            impacts = news_impacts.get(nid, [])
            normalized = info['normalized']

            if not impacts:
                # No impact data - classify as ambiguous
                templates[normalized] = {
                    'headline_example': info['headline'],
                    'direction': 'ambiguous',
                    'severity': 'small',
                    'category': 'FIN',
                    'confidence': 0.5,
                    'observed_count': templates.get(normalized, {}).get('observed_count', 0) + 1,
                    'avg_target_move': 0,
                    'avg_prob_change': 0,
                }
                continue

            # Find the biggest target move
            biggest = max(impacts, key=lambda x: abs(x.get('price_change', 0) or 0))
            price_move = biggest.get('price_change', 0) or 0
            prob_change = biggest.get('prob_change', 0) or 0
            deal_id = biggest.get('deal_id')

            # Determine direction from price movement
            if abs(price_move) < 0.05:
                direction = 'ambiguous'
            elif price_move > 0:
                direction = 'positive'
            else:
                direction = 'negative'

            # Determine severity from probability change magnitude
            abs_prob = abs(prob_change)
            if abs_prob >= 0.08:
                severity = 'large'
            elif abs_prob >= 0.03:
                severity = 'medium'
            else:
                severity = 'small'

            # Determine category from headline keywords
            category = self._guess_category(info['headline'])

            # Update or create template entry
            if normalized in templates:
                entry = templates[normalized]
                entry['observed_count'] += 1
                entry['avg_target_move'] = (
                    (entry['avg_target_move'] * (entry['observed_count'] - 1) + price_move)
                    / entry['observed_count'])
                entry['avg_prob_change'] = (
                    (entry['avg_prob_change'] * (entry['observed_count'] - 1) + prob_change)
                    / entry['observed_count'])
            else:
                templates[normalized] = {
                    'headline_example': info['headline'],
                    'direction': direction,
                    'severity': severity,
                    'category': category,
                    'confidence': 1.0 if abs(price_move) > 0.10 else 0.8,
                    'observed_count': 1,
                    'avg_target_move': price_move,
                    'avg_prob_change': prob_change,
                    'deal_pattern': deal_id,
                }

        # Save JSON
        json_path = os.path.join(self.output_dir, f'news_templates_{self.heat_label}.json')
        with open(json_path, 'w') as f:
            json.dump(templates, f, indent=2, sort_keys=True)
        print(f"    Saved {len(templates)} templates to {json_path}")

        # Also generate a Python snippet ready to paste into merger_arb_trader.py
        py_path = os.path.join(self.output_dir, f'news_templates_snippet_{self.heat_label}.py')
        with open(py_path, 'w') as f:
            f.write("# Auto-generated from practice session data\n")
            f.write("# Copy these entries into NEWS_TEMPLATES in merger_arb_trader.py\n\n")
            f.write("NEW_TEMPLATES = {\n")
            for norm, entry in sorted(templates.items()):
                d = entry['direction']
                s = entry['severity']
                cat = entry['category']
                conf = entry.get('confidence', 1.0)
                move = entry.get('avg_target_move', 0)
                prob = entry.get('avg_prob_change', 0)
                count = entry.get('observed_count', 1)
                f.write(f"    '{norm}': {{\n")
                f.write(f"        'direction': '{d}', 'severity': '{s}', "
                        f"'category': '{cat}', 'confidence': {conf},\n")
                f.write(f"        # seen {count}x, avg target move=${move:+.3f}, "
                        f"avg prob change={prob:+.4f}\n")
                f.write(f"    }},\n")
            f.write("}\n")
        print(f"    Saved Python snippet to {py_path}")

    def _guess_category(self, headline: str) -> str:
        """Guess the news category from headline keywords."""
        h = headline.upper()
        cat_scores = {'REG': 0, 'FIN': 0, 'SHR': 0, 'ALT': 0, 'PRC': 0}

        reg_kw = ['REGULAT', 'ANTITRUST', 'APPROVAL', 'FTC', 'DOJ', 'FDIC',
                   'COMMISSION', 'REVIEW', 'FILING', 'REMEDY', 'PHASE II',
                   'CLEARANCE', 'INVESTIGATION', 'FERC', 'OCC', 'SEC ',
                   'LEGAL', 'LAWSUIT', 'INJUNCTION', 'CLASS ACTION']
        fin_kw = ['FINANC', 'EARNINGS', 'REVENUE', 'CREDIT', 'DEBT', 'PROFIT',
                   'VALUATION', 'DOWNGRADE', 'UPGRADE', 'SYNERG', 'RETENTION',
                   'QUARTER', 'DIVIDEND', 'LIQUIDITY', 'REFINANC']
        shr_kw = ['SHAREHOLD', 'VOTE', 'PROXY', 'ACTIVIST', 'BOARD', 'DIRECTOR',
                   'INVESTOR', 'MANAGEMENT', 'SPECULATION', 'INSTITUTIONAL']
        alt_kw = ['COMPETING', 'RIVAL', 'COUNTER', 'BIDDER', 'UNSOLICITED',
                   'SWEETENED', 'WHITE KNIGHT', 'SUPERIOR']
        prc_kw = ['PRICE', 'PREMIUM', 'DISCOUNT', 'SPREAD', 'FAIR VALUE',
                   'BOOK VALUE']

        for kw in reg_kw:
            if kw in h:
                cat_scores['REG'] += 1
        for kw in fin_kw:
            if kw in h:
                cat_scores['FIN'] += 1
        for kw in shr_kw:
            if kw in h:
                cat_scores['SHR'] += 1
        for kw in alt_kw:
            if kw in h:
                cat_scores['ALT'] += 1
        for kw in prc_kw:
            if kw in h:
                cat_scores['PRC'] += 1

        best = max(cat_scores, key=cat_scores.get)
        return best if cat_scores[best] > 0 else 'FIN'

    def _generate_text_report(self):
        """Generate a human-readable session report."""
        print(f"  Building session_report_{self.heat_label}.txt ...")
        report_path = os.path.join(self.output_dir, f'session_report_{self.heat_label}.txt')
        c = self.conn.cursor()

        with open(report_path, 'w') as f:
            f.write("=" * 100 + "\n")
            f.write("PRACTICE SESSION REPORT\n")
            f.write(f"Generated: {datetime.now().isoformat()}\n")
            f.write(f"Database: {self.db_path}\n")
            f.write("=" * 100 + "\n\n")

            # NLV summary
            c.execute('SELECT MIN(nlv), MAX(nlv), '
                      'MIN(tick), MAX(tick) FROM ticks WHERE nlv > 0')
            row = c.fetchone()
            if row and row[0]:
                f.write(f"NLV: ${row[0]:,.2f} (min) -> ${row[1]:,.2f} (max)\n")
                f.write(f"Ticks: {row[2]} to {row[3]}\n\n")

            # News summary
            f.write("-" * 100 + "\n")
            f.write("NEWS ITEMS AND THEIR PRICE IMPACT\n")
            f.write("-" * 100 + "\n\n")

            c.execute('''
                SELECT n.news_id, n.tick, n.headline, n.headline_normalized,
                       GROUP_CONCAT(
                           ni.ticker || ':' ||
                           ROUND(ni.price_change, 3) || ':' ||
                           ROUND(ni.prob_change, 4),
                           ' | '
                       ) as impacts
                FROM news n
                LEFT JOIN news_impact ni ON n.news_id = ni.news_id AND ni.is_target = 1
                GROUP BY n.news_id
                ORDER BY n.tick
            ''')

            for row in c.fetchall():
                nid, tick, headline, normalized, impacts = row
                f.write(f"#{nid:3d} tick={tick:3d}: {headline}\n")
                f.write(f"     normalized: {normalized}\n")
                if impacts:
                    f.write(f"     impact: {impacts}\n")
                f.write("\n")

            # Word frequency analysis
            f.write("-" * 100 + "\n")
            f.write("WORD FREQUENCY IN NEWS HEADLINES\n")
            f.write("-" * 100 + "\n\n")

            word_counts = defaultdict(int)
            positive_words = defaultdict(int)
            negative_words = defaultdict(int)

            c.execute('SELECT news_id, headline FROM news')
            for nid, headline in c.fetchall():
                # Get the biggest target impact for this news
                c2 = self.conn.cursor()
                c2.execute('''SELECT price_change FROM news_impact
                             WHERE news_id=? AND is_target=1
                             ORDER BY abs(price_change) DESC LIMIT 1''', (nid,))
                impact_row = c2.fetchone()
                impact = impact_row[0] if impact_row else 0

                words = headline.upper().split()
                for word in words:
                    word = word.strip('.,!?;:()[]{}')
                    if len(word) < 3:
                        continue
                    word_counts[word] += 1
                    if impact > 0.05:
                        positive_words[word] += 1
                    elif impact < -0.05:
                        negative_words[word] += 1

            f.write("Top words in POSITIVE-impact news:\n")
            for word, count in sorted(positive_words.items(),
                                       key=lambda x: -x[1])[:40]:
                neg_count = negative_words.get(word, 0)
                ratio = count / (count + neg_count) if (count + neg_count) > 0 else 0
                f.write(f"  {word:25s} pos={count:2d} neg={neg_count:2d} "
                        f"ratio={ratio:.0%}\n")

            f.write("\nTop words in NEGATIVE-impact news:\n")
            for word, count in sorted(negative_words.items(),
                                       key=lambda x: -x[1])[:40]:
                pos_count = positive_words.get(word, 0)
                ratio = count / (count + pos_count) if (count + pos_count) > 0 else 0
                f.write(f"  {word:25s} neg={count:2d} pos={pos_count:2d} "
                        f"ratio={ratio:.0%}\n")

            # Deal probability trajectories
            f.write("\n" + "-" * 100 + "\n")
            f.write("DEAL PROBABILITY TRAJECTORIES\n")
            f.write("-" * 100 + "\n\n")

            for did in ['D1', 'D2', 'D3', 'D4', 'D5']:
                c.execute('''SELECT tick, market_implied_prob, deal_value,
                            target_price, acquirer_price, spread
                           FROM probabilities WHERE deal_id=?
                           ORDER BY tick''', (did,))
                rows = c.fetchall()
                if rows:
                    first = rows[0]
                    last = rows[-1]
                    min_p = min(r[1] for r in rows)
                    max_p = max(r[1] for r in rows)
                    f.write(f"{did}: {first[1]:.1%} -> {last[1]:.1%} "
                            f"(range {min_p:.1%}-{max_p:.1%})\n")

            # Orders summary
            f.write("\n" + "-" * 100 + "\n")
            f.write("ORDERS PLACED\n")
            f.write("-" * 100 + "\n\n")

            c.execute('''SELECT tick, ticker, action, quantity, price, status
                        FROM orders ORDER BY tick''')
            for row in c.fetchall():
                otick = row[0] if row[0] is not None else 0
                oticker = row[1] or '???'
                oaction = row[2] or '????'
                oquant = row[3] if row[3] is not None else 0
                oprice = row[4] if row[4] is not None else 0.0
                ostatus = row[5] or 'UNKNOWN'
                f.write(f"  tick={otick:3d}: {oaction:4s} {oquant:5d} {oticker:3s} "
                        f"@ ${oprice:8.2f} [{ostatus}]\n")

        print(f"    Saved to {report_path}")

    def _generate_excel_report(self):
        """Generate Excel workbook with all data tables."""
        try:
            import openpyxl
        except ImportError:
            print("    (openpyxl not installed - skipping Excel export)")
            print("    Install with: pip install openpyxl")
            return

        print(f"  Building session_data_{self.heat_label}.xlsx ...")
        xlsx_path = os.path.join(self.output_dir, f'session_data_{self.heat_label}.xlsx')
        wb = openpyxl.Workbook()

        c = self.conn.cursor()

        # Sheet 1: News + Impact
        ws = wb.active
        ws.title = "News"
        ws.append(['news_id', 'tick', 'headline', 'normalized', 'body',
                    'most_affected_target', 'target_move', 'target_move_pct',
                    'prob_before', 'prob_after', 'prob_change', 'deal_id'])

        c.execute('''
            SELECT n.news_id, n.tick, n.headline, n.headline_normalized, n.body,
                   ni.ticker, ni.price_change, ni.price_change_pct,
                   ni.prob_before, ni.prob_after, ni.prob_change, ni.deal_id
            FROM news n
            LEFT JOIN news_impact ni ON n.news_id = ni.news_id AND ni.is_target = 1
            ORDER BY n.tick, abs(ni.price_change) DESC
        ''')
        seen_news = set()
        for row in c.fetchall():
            if row[0] not in seen_news:
                seen_news.add(row[0])
                ws.append(list(row))

        # Sheet 2: Tick-level prices
        ws2 = wb.create_sheet("Prices")
        header = ['tick'] + ALL_TICKERS
        ws2.append(header)

        c.execute('SELECT DISTINCT tick FROM prices ORDER BY tick')
        ticks = [r[0] for r in c.fetchall()]
        for tick in ticks:
            row_data = [tick]
            for ticker in ALL_TICKERS:
                c.execute('SELECT mid FROM prices WHERE tick=? AND ticker=?',
                          (tick, ticker))
                r = c.fetchone()
                row_data.append(r[0] if r else '')
            ws2.append(row_data)

        # Sheet 3: Probabilities
        ws3 = wb.create_sheet("Probabilities")
        ws3.append(['tick', 'D1', 'D2', 'D3', 'D4', 'D5'])

        c.execute('SELECT DISTINCT tick FROM probabilities ORDER BY tick')
        ticks = [r[0] for r in c.fetchall()]
        for tick in ticks:
            row_data = [tick]
            for did in ['D1', 'D2', 'D3', 'D4', 'D5']:
                c.execute('SELECT market_implied_prob FROM probabilities '
                          'WHERE tick=? AND deal_id=?', (tick, did))
                r = c.fetchone()
                row_data.append(r[0] if r else '')
            ws3.append(row_data)

        # Sheet 4: NLV over time
        ws4 = wb.create_sheet("NLV")
        ws4.append(['tick', 'nlv', 'gross_position', 'net_position'])
        c.execute('SELECT tick, nlv, gross_position, net_position '
                  'FROM ticks ORDER BY tick')
        for row in c.fetchall():
            ws4.append(list(row))

        # Sheet 5: Orders
        ws5 = wb.create_sheet("Orders")
        ws5.append(['order_id', 'tick', 'ticker', 'type', 'action',
                     'quantity', 'price', 'status'])
        c.execute('SELECT * FROM orders ORDER BY tick')
        for row in c.fetchall():
            ws5.append(list(row))

        # Sheet 6: Word Bank
        ws6 = wb.create_sheet("Word Bank")
        ws6.append(['word', 'total_count', 'positive_news_count',
                     'negative_news_count', 'neutral_news_count',
                     'positive_ratio', 'signal_strength'])

        word_data = defaultdict(lambda: {'total': 0, 'pos': 0, 'neg': 0, 'neu': 0})
        c.execute('SELECT n.news_id, n.headline FROM news n')
        for nid, headline in c.fetchall():
            c2 = self.conn.cursor()
            c2.execute('''SELECT price_change FROM news_impact
                         WHERE news_id=? AND is_target=1
                         ORDER BY abs(price_change) DESC LIMIT 1''', (nid,))
            impact_row = c2.fetchone()
            impact = impact_row[0] if impact_row else 0

            words = headline.upper().split()
            for word in words:
                word = word.strip('.,!?;:()[]{}')
                if len(word) < 3:
                    continue
                word_data[word]['total'] += 1
                if impact > 0.05:
                    word_data[word]['pos'] += 1
                elif impact < -0.05:
                    word_data[word]['neg'] += 1
                else:
                    word_data[word]['neu'] += 1

        for word, counts in sorted(word_data.items(),
                                    key=lambda x: -(x[1]['pos'] + x[1]['neg'])):
            total = counts['total']
            pos = counts['pos']
            neg = counts['neg']
            neu = counts['neu']
            directional = pos + neg
            pos_ratio = pos / directional if directional > 0 else 0.5
            # Signal strength = how often this word appears in directional news
            strength = directional / total if total > 0 else 0
            ws6.append([word, total, pos, neg, neu,
                        round(pos_ratio, 3), round(strength, 3)])

        wb.save(xlsx_path)
        print(f"    Saved to {xlsx_path}")


# =============================================================================
# ENTRY POINT
# =============================================================================
def main():
    parser = argparse.ArgumentParser(description='Practice Session Logger')
    parser.add_argument('--api-key', default=DEFAULT_API_KEY,
                        help=f'RIT API key (default: {DEFAULT_API_KEY})')
    parser.add_argument('--port', type=int, default=10000,
                        help='RIT API port (default: 10000)')
    parser.add_argument('--output', default=None,
                        help='Output directory (default: ../data/practice_logs/)')
    args = parser.parse_args()

    base_url = f'http://localhost:{args.port}/v1'

    if args.output:
        output_dir = args.output
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, '..', 'data', 'practice_logs')

    logger = PracticeLogger(args.api_key, base_url, output_dir)
    logger.run()


if __name__ == '__main__':
    main()
